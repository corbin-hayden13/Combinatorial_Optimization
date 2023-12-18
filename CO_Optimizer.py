import openpyxl as pxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


grain_flow_path = "EnviroSpec_Vision/2023_Post_Low_Carbon_grain_flow_impact_calcs_ver_001_2023-10-31.xlsx"
wheat_supply_path = "EnviroSpec_Vision/Indigo 2023_Wheat_Airly_Summary.xlsx"


def scaled_sigmoid(score, min_bound=-10, max_bound=10, k_steepness=0.0001, x0_sigmoid_midpoint=0):
    """
    k_steepness: Closer to 0, more gradual approach to min / max is to accommodate for very high values
    """
    sigmoid = max_bound / (1 + np.exp(-k_steepness * (score - x0_sigmoid_midpoint)))
    return (np.abs(max_bound) + np.abs(min_bound)) * (sigmoid / max_bound) + min_bound


def evaluate_individual(individual, actual_values):
    ret_dict = {}

    for index, key in individual:
        try:
            ret_dict[key] += actual_values[index]
        except KeyError:
            ret_dict[key] = actual_values[index]

    return ret_dict


def precision_round(value, precision=5):
    return float(f"{value:.{precision}f}")


def min_field_count_make_individual(row_vals, max_lots, min_field_count):
    lots = np.repeat(np.arange(max_lots), min_field_count)

    # Distribute any remaining ingredients
    remaining_fields = len(row_vals) % min_field_count
    if remaining_fields > 0:
        extra_lots = np.random.choice(max_lots, remaining_fields, replace=False)
        lots = np.concatenate([lots, extra_lots])

    # Shuffle the ingredients and pair them with meals
    np.random.shuffle(lots)
    fields = np.random.permutation(len(row_vals))

    # Create tuples of (ingredient_index, meal_number)
    individual = list(zip(fields, lots))

    return individual


def make_individual(row_vals, max_lots):
    indices = np.arange(len(row_vals))
    random_integers = np.random.randint(0, max_lots, size=len(row_vals))
    return list(zip(indices, random_integers))


def make_population(row_vals, max_lots, min_field_count=-1, pop_size=100):
    if min_field_count == -1: return [make_individual(row_vals, max_lots) for _ in range(pop_size)]
    else: return [min_field_count_make_individual(row_vals, max_lots, min_field_count) for _ in range(pop_size)]


def clone_individual(individual, pop_size=100):
    return [individual for _ in range(pop_size)]


def fitness(row_vals, individual, target_val):
    """
    What's being rewarded?
     - euclidean norm of combos
     - more lots than fewer lots
     - more negative scores than positive scores
     - variance of sums of lots
    """
    gen_mod = 1e-6
    fitness_dict = {}
    for index, label in individual:
        try:
            fitness_dict[label] += row_vals[index]
        except KeyError:
            fitness_dict[label] = row_vals[index]

    combo_scores = [target_val - fitness_dict[key] for key in fitness_dict.keys()]
    negative_mod = sum([-gen_mod if score <= 0 else gen_mod for score in combo_scores])
    euclidean_norm = np.linalg.norm(combo_scores)

    sigmoid_variance_range = 1.99

    return (2 * (1 / (euclidean_norm + gen_mod))) + \
           (1 * (1 / (len(row_vals) / len(fitness_dict.keys()) + gen_mod))) + \
           (-1.45 * negative_mod) - (scaled_sigmoid(np.var(combo_scores), min_bound=-sigmoid_variance_range,
                                                    max_bound=sigmoid_variance_range, k_steepness=0.000001))


def fit_population(row_vals, population, target_val):
    return [fitness(row_vals, individual, target_val) for individual in population]


def vector_from_dataframe_column(dataframe, header_row=1, col=0):
    try:
        return np.array([dataframe.cell(header_row + a, col).value for a in range(1, dataframe.max_row + 1)
                         if type(dataframe.cell(header_row + a, col).value) != str
                         and dataframe.cell(header_row + a, col).value is not None])
    except AttributeError:
        return dataframe


def load_data(file_name, workbook=None, data_only=True, print_sheet_names=False):
    dataframe = pxl.load_workbook(file_name, data_only=data_only)
    if print_sheet_names: print(dataframe.sheetnames)
    if workbook is None:
        return dataframe.active
    else:
        return dataframe[workbook]


def uniform_test_data(N, min_val, max_val, percent_negative=0.5, shuffle=True):
    if min_val > 0 or max_val < 0: return np.linspace(min_val, max_val, N)

    else:
        num_negative = int(N * percent_negative)
        num_positive = N - num_negative

        ret_list = np.hstack((np.linspace(min_val, -1, num_negative), np.linspace(1, max_val, num_positive)))
        if shuffle: np.random.shuffle(ret_list)

        return ret_list


class Optimizer:
    def __init__(self, algorithm="ga", init_dataframe=None, header_row=1, max_column=10):
        self.dataframe = init_dataframe
        self.header_row = header_row
        self.max_column = max_column
        self.evaluated_column = None

        if algorithm == "ga": self.best_individual = None

    def vectorize_column(self, col=0):
        if self.dataframe is None:
            print("Please call \"import_data()\" before \"optimize_for()\" to load a dataset into the optimizer")
            return
        else:
            return vector_from_dataframe_column(self.dataframe, col=col)

    def import_data(self, file_name, header_row=1, workbook=None, data_only=True, print_sheet_names=False):
        self.dataframe = load_data(file_name, workbook=workbook, data_only=data_only,
                                   print_sheet_names=print_sheet_names)
        self.header_row = header_row
        return self

    def load_test_data(self, test_size=32, min_bound=-10000, max_bound=10000, percent_negative=0.5, shuffle=True):
        self.dataframe = uniform_test_data(test_size, min_bound, max_bound, percent_negative=percent_negative,
                                           shuffle=shuffle).T

        self.header_row = -1
        return self

    def individual_to_dict(self, individual, sort_keys=False):
        ret_dict = {}
        for index, key in individual:
            try:
                ret_dict[key].append(index)
            except KeyError:
                ret_dict[key] = [index]

        if sort_keys:
            return {key: ret_dict[key] for key in sorted(list(ret_dict.keys()))}
        else: return ret_dict

    def graph_data(self, col=0):
        if self.dataframe is None:
            print("Please call \"import_data()\" before \"optimize_for()\" to load a dataset into the optimizer")
            return
        else:
            sorted_values = np.abs(sorted(vector_from_dataframe_column(self.dataframe, self.header_row, col=col)))
            plt.plot(sorted_values)
            plt.show()

    def to_csv(self, file_name, start_column=1, replace_column="Lot #"):
        csv_dict = {}
        if self.dataframe is None:
            print("Please call \"import_data()\" then \"optimize_for()\" to load a dataset into and train the optimizer")
            return
        if self.best_individual is None and self.dataframe is not None:
            print("Please call \"optimize_for()\" to train the optimizer")
            return

        for col_num in range(start_column, self.max_column + 1):
            csv_dict[self.dataframe.cell(self.header_row, col_num).value] = []

        individual_dict = self.individual_to_dict(self.best_individual, sort_keys=True)
        for lot_num in individual_dict.keys():
            for row_ind in individual_dict[lot_num]:
                csv_keys_list = list(csv_dict.keys())
                for a in range(len(list(csv_keys_list))):
                    if csv_keys_list[a] == replace_column:
                        csv_dict[csv_keys_list[a]].append(lot_num + 2023001)
                    else:
                        csv_dict[csv_keys_list[a]].append(self.dataframe.cell(row_ind + self.header_row + 1, a + 1).value)

        new_dataframe = pd.DataFrame(csv_dict)
        new_dataframe.to_csv(file_name)